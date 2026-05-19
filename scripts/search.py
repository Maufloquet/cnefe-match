"""Busca a base de busca contra o índice cnefe_ba em 3 camadas.

A camada 1 exige município + CEP + número, com logradouro fuzzy.
A 2 relaxa o número. A 3 relaxa o CEP (vira prefixo de 5 dígitos).
A primeira camada que retornar hit com score acima do limite vence.

Cada endereço sai com um match_status: encontrado, ambiguo ou
nao_encontrado. Ambiguidade é detectada comparando os 2 hits de
maior score do ES — se o segundo está perto demais do primeiro
(razão >= AMBIG_RATIO) e os setores são distintos, marcamos.

Saída em data/processed/resultado.csv com 4 colunas novas:
setor_censitario_encontrado, match_status, match_layer e match_score.
"""

import sys
from pathlib import Path

import pandas as pd
from elasticsearch import Elasticsearch

sys.path.insert(0, str(Path(__file__).parent))
from normalize import normalize_row

ES_URL = "http://localhost:9200"
INDEX = "cnefe_ba"

XLSX_IN = Path(__file__).parent.parent / "data" / "raw" / "base_busca.xlsx"
CSV_OUT = Path(__file__).parent.parent / "data" / "processed" / "resultado.csv"
XLSX_OUT = Path(__file__).parent.parent / "data" / "processed" / "resultado.xlsx"

# scores mínimos por camada — empíricos, ajustar depois da 1a rodada
MIN_SCORE_LAYER1 = 5.0
MIN_SCORE_LAYER2 = 3.0
MIN_SCORE_LAYER3 = 1.0

# Se o segundo melhor hit tiver score >= AMBIG_RATIO * primeiro,
# consideramos o resultado ambíguo (a menos que os setores coincidam).
AMBIG_RATIO = 0.85


def query_layer1(n):
    must = [
        {"term": {"cod_municipio_6": n["municipio_6"]}},
        {"term": {"cep": n["cep"]}},
    ]
    if n["numero"]:
        must.append({"term": {"numero": n["numero"]}})
    if n["logradouro"]:
        must.append({
            "match": {
                "logradouro_full": {
                    "query": n["logradouro"],
                    "fuzziness": "AUTO",
                }
            }
        })
    return {"bool": {"must": must}}


def query_layer2(n):
    must = [
        {"term": {"cod_municipio_6": n["municipio_6"]}},
        {"term": {"cep": n["cep"]}},
    ]
    should = []
    if n["logradouro"]:
        should.append({"match": {"logradouro_full": {"query": n["logradouro"], "fuzziness": "AUTO"}}})
    if n["bairro"]:
        should.append({"match": {"bairro": {"query": n["bairro"], "fuzziness": "AUTO"}}})
    q = {"bool": {"must": must}}
    if should:
        q["bool"]["should"] = should
        q["bool"]["minimum_should_match"] = 1
    return q


def query_layer3(n):
    # CEP no Brasil: 5 primeiros = região/bairro, 3 últimos = face de quadra.
    # prefixo dos 5 sobrevive a erros de digitação no fim.
    must = [
        {"term": {"cod_municipio_6": n["municipio_6"]}},
    ]
    should = []
    if n["logradouro"]:
        should.append({"match": {"logradouro_full": {"query": n["logradouro"], "fuzziness": "AUTO", "boost": 2.0}}})
    if n["bairro"]:
        should.append({"match": {"bairro": {"query": n["bairro"], "fuzziness": "AUTO"}}})
    if n["cep"]:
        should.append({"prefix": {"cep": {"value": n["cep"][:5], "boost": 1.5}}})
    q = {"bool": {"must": must}}
    if should:
        q["bool"]["should"] = should
        q["bool"]["minimum_should_match"] = 1
    return q


def is_ambiguous(hits):
    """Top 2 hits muito proximos em score e com setores diferentes."""
    if len(hits) < 2:
        return False
    top1, top2 = hits[0], hits[1]
    s1, s2 = top1["_score"], top2["_score"]
    if s1 <= 0:
        return False
    if top1["_source"]["cod_setor"] == top2["_source"]["cod_setor"]:
        return False
    return (s2 / s1) >= AMBIG_RATIO


def search_one(es, n):
    """Tenta camada 1 -> 2 -> 3 e devolve (cod_setor, status, layer, score).

    status: 'encontrado', 'ambiguo' ou 'nao_encontrado'.
    """
    if not n["municipio_6"]:
        return (None, "nao_encontrado", "none", 0.0)

    layers = [
        (1, query_layer1, MIN_SCORE_LAYER1),
        (2, query_layer2, MIN_SCORE_LAYER2),
        (3, query_layer3, MIN_SCORE_LAYER3),
    ]
    for layer_num, builder, min_score in layers:
        try:
            resp = es.search(
                index=INDEX,
                query=builder(n),
                size=5,
                _source=["cod_setor"],
            )
            hits = resp["hits"]["hits"]
            if not hits:
                continue
            score = hits[0]["_score"]
            if score < min_score:
                continue
            status = "ambiguo" if is_ambiguous(hits) else "encontrado"
            return (hits[0]["_source"]["cod_setor"], status, layer_num, score)
        except Exception as e:
            print(f"  erro camada {layer_num}: {e}")
    return (None, "nao_encontrado", "none", 0.0)


def read_input(path: Path) -> pd.DataFrame:
    return pd.read_excel(path, dtype=str)


def process(df: pd.DataFrame, es: Elasticsearch) -> pd.DataFrame:
    setores, statuses, layers, scores = [], [], [], []
    for i, row in df.iterrows():
        n = normalize_row(row.to_dict())
        setor, status, layer, score = search_one(es, n)
        setores.append(setor)
        statuses.append(status)
        layers.append(layer)
        scores.append(round(score, 2))
        mark = {"encontrado": "ok ", "ambiguo": "amb", "nao_encontrado": "---"}[status]
        print(f"  {i+1:3d}. [{mark}] layer={layer} score={score:5.2f} -> {setor}")

    out = df.copy()
    out["setor_censitario_encontrado"] = setores
    out["match_status"] = statuses
    out["match_layer"] = layers
    out["match_score"] = scores
    return out


COL_WIDTHS = {
    "consulta_logradouro": 38,
    "consulta_numero": 14,
    "consulta_bairro": 26,
    "consulta_cep": 12,
    "consulta_complemento": 18,
    "consulta_municipio": 14,
    "setor_censitario_encontrado": 24,
    "match_status": 16,
    "match_layer": 12,
    "match_score": 12,
}


def write_output(df: pd.DataFrame, csv_path: Path, xlsx_path: Path) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(csv_path, index=False)
    df.to_excel(xlsx_path, index=False, sheet_name="resultado")

    # Ajustes cosmeticos no xlsx pra abrir legivel sem mexer no Excel
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = load_workbook(xlsx_path)
    ws = wb["resultado"]
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True)
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    for idx, col_name in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = COL_WIDTHS.get(col_name, 18)

    # cores por match_status pra facilitar inspecao visual
    status_fills = {
        "encontrado": PatternFill(start_color="D8F0D8", end_color="D8F0D8", fill_type="solid"),
        "ambiguo": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "nao_encontrado": PatternFill(start_color="F8CECC", end_color="F8CECC", fill_type="solid"),
    }
    status_col = list(df.columns).index("match_status") + 1
    for row in range(2, ws.max_row + 1):
        fill = status_fills.get(ws.cell(row=row, column=status_col).value)
        if fill:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    wb.save(xlsx_path)


def print_summary(df: pd.DataFrame) -> None:
    total = len(df)
    counts = df["match_status"].value_counts().to_dict()
    layer_counts = df["match_layer"].value_counts().to_dict()
    enc = counts.get("encontrado", 0)
    amb = counts.get("ambiguo", 0)
    neg = counts.get("nao_encontrado", 0)
    print(f"\nResumo:")
    print(f"  total:           {total}")
    print(f"  encontrados:     {enc} ({100 * enc / total:.1f}%)")
    print(f"  ambiguos:        {amb}")
    print(f"  nao encontrados: {neg}")
    for k in [1, 2, 3, "1", "2", "3", "none"]:
        if k in layer_counts:
            print(f"  camada {k}: {layer_counts[k]}")


def main():
    es = Elasticsearch(ES_URL, request_timeout=60)
    if not es.ping():
        print("ES nao responde")
        sys.exit(1)

    df = read_input(XLSX_IN)
    print(f"Processando {len(df)} enderecos...\n")
    out = process(df, es)
    write_output(out, CSV_OUT, XLSX_OUT)
    print_summary(out)
    print(f"\nGravado em {CSV_OUT}")
    print(f"          {XLSX_OUT}")


if __name__ == "__main__":
    main()
